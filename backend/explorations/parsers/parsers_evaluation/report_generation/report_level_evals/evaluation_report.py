"""
Generate comprehensive evaluation report with metrics, Excel dashboard, and AI analysis.

This script analyzes the evaluation results from report-level evaluations,
computing metrics, generating Excel dashboards with charts, and producing
AI-generated insights about pipeline performance.
"""

import asyncio
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from dotenv import load_dotenv
from google import genai
from google.genai.types import GenerateContentConfig, ThinkingConfig
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

load_dotenv()

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))


def load_evaluation_results(eval_dir: Path) -> list[dict[str, Any]]:
    """Load all evaluation results from JSON files in the eval directory."""
    all_results = []
    for json_file in eval_dir.glob("*.json"):
        with open(json_file, encoding="utf-8") as f:
            results = json.load(f)
            all_results.extend(results)
    return all_results


def compute_global_metrics(results: list[dict[str, Any]]) -> dict[str, Any]:
    """Compute overall performance metrics across all evaluations."""
    total = len(results)
    if total == 0:
        return {}

    verdict_counts = {
        "correct": sum(1 for r in results if r["verdict"] == "CORRECT"),
        "partial": sum(1 for r in results if r["verdict"] == "PARTIAL"),
        "hallucination": sum(1 for r in results if r["verdict"] == "HALLUCINATION"),
        "missed": sum(1 for r in results if r["verdict"] == "MISSED"),
    }

    avg_confidence = sum(r.get("confidence", 0.0) for r in results) / total

    return {
        "total_evaluations": total,
        "correct_count": verdict_counts["correct"],
        "partial_count": verdict_counts["partial"],
        "hallucination_count": verdict_counts["hallucination"],
        "missed_count": verdict_counts["missed"],
        "correct_rate": verdict_counts["correct"] / total,
        "partial_rate": verdict_counts["partial"] / total,
        "hallucination_rate": verdict_counts["hallucination"] / total,
        "missed_rate": verdict_counts["missed"] / total,
        "avg_confidence": avg_confidence,
        "success_rate": (verdict_counts["correct"] + verdict_counts["partial"]) / total,
    }


def compute_pipeline_metrics(results: list[dict[str, Any]]) -> pd.DataFrame:
    """Compute performance metrics grouped by pipeline."""
    df = pd.DataFrame(results)

    pipeline_stats = (
        df.groupby("pipeline_name")
        .agg(
            {
                "verdict": [
                    ("total", "count"),
                    ("correct", lambda x: (x == "CORRECT").sum()),
                    ("partial", lambda x: (x == "PARTIAL").sum()),
                    ("hallucination", lambda x: (x == "HALLUCINATION").sum()),
                    ("missed", lambda x: (x == "MISSED").sum()),
                ],
                "confidence": "mean",
            }
        )
        .round(4)
    )

    pipeline_stats.columns = ["_".join(col).strip("_") for col in pipeline_stats.columns]

    # Calculate rates
    pipeline_stats["correct_rate"] = (
        pipeline_stats["verdict_correct"] / pipeline_stats["verdict_total"]
    ).round(4)
    pipeline_stats["partial_rate"] = (
        pipeline_stats["verdict_partial"] / pipeline_stats["verdict_total"]
    ).round(4)
    pipeline_stats["hallucination_rate"] = (
        pipeline_stats["verdict_hallucination"] / pipeline_stats["verdict_total"]
    ).round(4)
    pipeline_stats["missed_rate"] = (
        pipeline_stats["verdict_missed"] / pipeline_stats["verdict_total"]
    ).round(4)
    pipeline_stats["success_rate"] = (
        (pipeline_stats["verdict_correct"] + pipeline_stats["verdict_partial"])
        / pipeline_stats["verdict_total"]
    ).round(4)

    return pipeline_stats.reset_index()


def create_detailed_results_df(results: list[dict[str, Any]]) -> pd.DataFrame:
    """Create a detailed DataFrame with all evaluation results."""
    rows: list[dict[str, Any]] = []
    for r in results:
        rows.append(
            {
                "pipeline": r["pipeline_name"],
                "report_file": r.get("report_file", ""),
                "category_path": r.get("category_path", ""),
                "expected_winner": r.get("expected_winner", ""),
                "question": r["question"],
                "verdict": r["verdict"],
                "confidence": r.get("confidence", 0.0),
                "ground_truth": r["ground_truth"],
                "llm_answer": r["llm_answer"],
                "explanation": r["explanation"],
                "missing_info": "; ".join(r.get("missing_information", [])),
            }
        )

    return pd.DataFrame(rows)


def compute_category_metrics(detailed_results: pd.DataFrame) -> pd.DataFrame:
    """Compute performance metrics grouped by category."""
    if detailed_results.empty:
        return pd.DataFrame()

    df = detailed_results.copy()
    group = df.groupby("category_path")
    stats = (
        group.agg(
            {
                "verdict": [
                    ("total", "count"),
                    ("correct", lambda x: (x == "CORRECT").sum()),
                    ("partial", lambda x: (x == "PARTIAL").sum()),
                    ("hallucination", lambda x: (x == "HALLUCINATION").sum()),
                    ("missed", lambda x: (x == "MISSED").sum()),
                ],
                "confidence": "mean",
            }
        )
        .round(4)
    )

    stats.columns = ["_".join(col).strip("_") for col in stats.columns]
    stats["correct_rate"] = (stats["verdict_correct"] / stats["verdict_total"]).round(4)
    stats["partial_rate"] = (stats["verdict_partial"] / stats["verdict_total"]).round(4)
    stats["hallucination_rate"] = (stats["verdict_hallucination"] / stats["verdict_total"]).round(4)
    stats["missed_rate"] = (stats["verdict_missed"] / stats["verdict_total"]).round(4)
    stats["success_rate"] = (
        (stats["verdict_correct"] + stats["verdict_partial"]) / stats["verdict_total"]
    ).round(4)

    return stats.reset_index()


def compute_pipeline_category_metrics(detailed_results: pd.DataFrame) -> pd.DataFrame:
    """Compute performance metrics grouped by pipeline and category."""
    if detailed_results.empty:
        return pd.DataFrame()

    df = detailed_results.copy()
    group = df.groupby(["pipeline", "category_path"])
    stats = group.agg(
        {
            "verdict": [
                ("total", "count"),
                ("correct", lambda x: (x == "CORRECT").sum()),
                ("partial", lambda x: (x == "PARTIAL").sum()),
                ("hallucination", lambda x: (x == "HALLUCINATION").sum()),
                ("missed", lambda x: (x == "MISSED").sum()),
            ],
        }
    )
    stats.columns = ["_".join(col).strip("_") for col in stats.columns]
    stats["success_rate"] = (
        (stats["verdict_correct"] + stats["verdict_partial"]) / stats["verdict_total"]
    ).round(4)
    return stats.reset_index()


def write_excel_with_charts(
    global_metrics: dict[str, Any],
    pipeline_metrics: pd.DataFrame,
    detailed_results: pd.DataFrame,
    category_metrics: pd.DataFrame,
    pipeline_category_metrics: pd.DataFrame,
    output_file: Path,
):
    """Write comprehensive Excel report with multiple sheets and charts."""
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Sheet 1: Global Summary
        global_df = pd.DataFrame([global_metrics])
        global_df.to_excel(writer, sheet_name="Global Summary", index=False)

        # Sheet 2: Pipeline Comparison
        pipeline_metrics.to_excel(writer, sheet_name="Pipeline Comparison", index=False)

        # Sheet 3: Detailed Results
        detailed_results.to_excel(writer, sheet_name="Detailed Results", index=False)

        # Sheet 4: Verdict Distribution by Pipeline
        verdict_pivot = detailed_results.pivot_table(
            index="pipeline", columns="verdict", aggfunc="size", fill_value=0
        )
        verdict_pivot.to_excel(writer, sheet_name="Verdict Distribution")

        # Sheet 5: Category Comparison
        if not category_metrics.empty:
            category_metrics.to_excel(writer, sheet_name="Category Comparison", index=False)

        # Sheet 6: Pipeline x Category (success rate pivot)
        if not pipeline_category_metrics.empty:
            pivot = pipeline_category_metrics.pivot(
                index="pipeline", columns="category_path", values="success_rate"
            ).reset_index()
            pivot.to_excel(writer, sheet_name="Pipeline x Category", index=False)

    # Add charts using openpyxl
    wb = load_workbook(output_file)

    # Add pie chart to Global Summary
    add_global_pie_chart(wb, global_metrics)

    # Add bar chart to Pipeline Comparison
    add_pipeline_bar_chart(wb, pipeline_metrics)

    # Add bar chart to Category Comparison
    if "Category Comparison" in wb.sheetnames and not category_metrics.empty:
        add_category_bar_chart(wb, category_metrics)

    # Add grouped bar chart to Pipeline x Category
    if "Pipeline x Category" in wb.sheetnames and not pipeline_category_metrics.empty:
        add_pipeline_category_grouped_bar_chart(wb)

    wb.save(output_file)


def add_global_pie_chart(wb, global_metrics: dict[str, Any]):
    """Add a pie chart showing overall verdict distribution."""
    ws = wb["Global Summary"]

    # Create a small data table for the chart
    chart_data = [
        ["Verdict", "Count"],
        ["Correct", global_metrics["correct_count"]],
        ["Partial", global_metrics["partial_count"]],
        ["Hallucination", global_metrics["hallucination_count"]],
        ["Missed", global_metrics["missed_count"]],
    ]

    start_row = ws.max_row + 3
    for i, row_data in enumerate(chart_data, start=start_row):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)

    # Create pie chart
    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + 4)
    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + 4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Overall Verdict Distribution"

    ws.add_chart(pie, f"E{start_row}")


def add_pipeline_bar_chart(wb, pipeline_metrics: pd.DataFrame):
    """Add a bar chart comparing success rates across pipelines."""
    ws = wb["Pipeline Comparison"]

    num_pipelines = len(pipeline_metrics)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Success Rate by Pipeline"
    chart.y_axis.title = "Success Rate"
    chart.x_axis.title = "Pipeline"

    data = Reference(ws, min_col=ws.max_column, min_row=1, max_row=num_pipelines + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=num_pipelines + 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    ws.add_chart(chart, f"A{num_pipelines + 5}")


def add_category_bar_chart(wb, category_metrics: pd.DataFrame):
    """Add a bar chart comparing success rates across categories."""
    ws = wb["Category Comparison"]

    num_categories = len(category_metrics)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Success Rate by Category"
    chart.y_axis.title = "Success Rate"
    chart.x_axis.title = "Category"

    data = Reference(ws, min_col=ws.max_column, min_row=1, max_row=num_categories + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=num_categories + 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    ws.add_chart(chart, f"A{num_categories + 5}")


def add_pipeline_category_grouped_bar_chart(wb):
    """Add a grouped bar chart showing success rates for each category per pipeline."""
    ws = wb["Pipeline x Category"]

    num_rows = ws.max_row
    num_cols = ws.max_column

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.overlap = 0
    chart.title = "Success Rate by Pipeline and Category"
    chart.y_axis.title = "Success Rate"
    chart.x_axis.title = "Pipeline"

    data = Reference(ws, min_col=2, min_row=1, max_col=num_cols, max_row=num_rows)
    cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    ws.add_chart(chart, f"A{num_rows + 3}")


async def generate_ai_report(
    global_metrics: dict[str, Any],
    pipeline_metrics: pd.DataFrame,
    detailed_results: pd.DataFrame,
    category_metrics: pd.DataFrame,
    pipeline_category_metrics: pd.DataFrame,
) -> str:
    """Generate comprehensive analysis report using Gemini with extended thinking (French)."""

    prompt = f"""
Tu es un expert data scientist analysant la performance de différents pipelines de génération de rapports de comparaison d'assurance.

## Méthodologie d'évaluation (description factuelle)

Nous avons évalué plusieurs pipelines en:
1. Générant des rapports de comparaison entre deux contrats d'assurance (ProBTP vs AXA)
2. Posant des questions ciblées sur les différences de couverture par catégorie
3. Comparant les réponses extraites des rapports au jeu de vérité terrain
4. Attribuant un verdict selon 4 catégories:
   - CORRECT: Concordance parfaite avec la vérité terrain
   - PARTIAL: Bonne direction mais informations clés manquantes
   - HALLUCINATION: Informations incorrectes ou non étayées
   - MISSED: Indique "non disponible" alors que l'information existe

## Pipelines évalués

- **two_phase**: Pipeline en deux phases avec extraction puis assemblage
- **taxonomy_first**: Pipeline basé taxonomie (extraction de structure puis valeurs)
- **baseline**: Pipeline baseline simple
- **project_merge**: Pipeline avec projection puis fusion

## Indicateurs globaux

{json.dumps(global_metrics, indent=2, ensure_ascii=False)}

## Comparaison des pipelines

{pipeline_metrics.to_string()}

## Performance par catégorie

{category_metrics.to_string() if not category_metrics.empty else "Aucune donnée par catégorie"}

## Pipeline x Catégorie (taux de succès)

{pipeline_category_metrics.pivot(index='pipeline', columns='category_path', values='success_rate').to_string() if not pipeline_category_metrics.empty else "Aucune donnée pipeline x catégorie"}

## Exemples de résultats détaillés (10 premiers)

{detailed_results.head(10).to_string()}

## Ta mission

Rédige un rapport d'analyse structuré en français, incluant:

1. Résumé exécutif: constats clés et recommandations (2-3 paragraphes)

2. Description de la méthodologie: reformulation neutre et factuelle

3. Analyse de performance:
   - Comparaison objective des pipelines
   - Points forts et faiblesses de chaque pipeline
   - Corrélation entre score de confiance et verdict

4. Analyse des causes:
   - Hypothèses sur les écarts de performance entre pipelines
   - Catégories les plus difficiles à évaluer correctement

5. Recommandations:
   - Pipeline(s) à privilégier et justification
   - Pistes d'amélioration pour chaque pipeline
   - Suggestions pour améliorer l'évaluation

6. Analyse par type de verdict:
   - PARTIAL: informations manquantes typiques
   - MISSED: angles morts systématiques éventuels
   - HALLUCINATION: sources fréquentes d'erreurs

7. Aperçus par catégorie:
   - Variabilité entre catégories et explications plausibles

Utilise un style clair, structuré en markdown, avec des listes lorsque pertinent.
"""

    client = genai.Client(
        vertexai=True,
        project="probtp-poc-prod",
        location="global",
    )

    try:
        response = await client.aio.models.generate_content(
            model="gemini-2.5-flash-preview-09-2025",
            contents=prompt,
            config=GenerateContentConfig(
                thinking_config=ThinkingConfig(thinking_budget=8192)  # Extended reasoning budget
            ),
        )

        report = getattr(response, "text", "").strip()
        return report

    finally:
        if hasattr(client, "close"):
            if asyncio.iscoroutinefunction(client.close):
                await client.close()
            else:
                client.close()


async def generate_full_report():
    """Main function to generate complete evaluation report."""
    print("Chargement des résultats d'évaluation...")
    eval_dir = Path(__file__).parent / "output" / "evals"

    if not eval_dir.exists():
        print(f"Error: Evaluation directory not found at {eval_dir}")
        print("Please run evaluate_reports.py first")
        return

    results = load_evaluation_results(eval_dir)
    print(f"{len(results)} résultats chargés")

    if not results:
        print("No results found. Please run evaluate_reports.py first.")
        return

    print("\nCalcul des métriques...")
    global_metrics = compute_global_metrics(results)
    pipeline_metrics = compute_pipeline_metrics(results)
    detailed_results = create_detailed_results_df(results)
    category_metrics = compute_category_metrics(detailed_results)
    pipeline_category_metrics = compute_pipeline_category_metrics(detailed_results)

    print("\nMétriques globales:")
    print(json.dumps(global_metrics, indent=2))

    print("\nMétriques par pipeline:")
    print(pipeline_metrics.to_string())

    if not category_metrics.empty:
        print("\nMétriques par catégorie:")
        print(category_metrics.to_string())

    if not pipeline_category_metrics.empty:
        print("\nPipeline x Catégorie (taux de succès):")
        print(
            pipeline_category_metrics.pivot(
                index='pipeline', columns='category_path', values='success_rate'
            ).to_string()
        )

    # Generate Excel report
    output_dir = Path(__file__).parent / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = output_dir / f"report_evaluation_{timestamp}.xlsx"

    print(f"\nGénération du rapport Excel: {excel_file}")
    write_excel_with_charts(
        global_metrics,
        pipeline_metrics,
        detailed_results,
        category_metrics,
        pipeline_category_metrics,
        excel_file,
    )
    print(f"✓ Rapport Excel enregistré: {excel_file}")

    # Generate AI analysis report
    print("\nGénération du rapport d'analyse IA (cela peut prendre une minute)...")
    ai_report = await generate_ai_report(
        global_metrics,
        pipeline_metrics,
        detailed_results,
        category_metrics,
        pipeline_category_metrics,
    )

    report_file = output_dir / f"ai_analysis_report_{timestamp}.md"
    with open(report_file, "w", encoding="utf-8") as f:
        f.write("# Rapport d'évaluation des pipelines de génération de rapports\n\n")
        f.write(f"**Généré :** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write("---\n\n")
        f.write(ai_report)

    print(f"✓ Rapport IA enregistré: {report_file}")

    return excel_file, report_file


if __name__ == "__main__":
    excel_path, report_path = asyncio.run(generate_full_report())
    print(f"\n{'=' * 60}")
    print("Génération du rapport terminée !")
    print(f"Tableau de bord Excel: {excel_path}")
    print(f"Analyse IA: {report_path}")
    print(f"{'=' * 60}")
