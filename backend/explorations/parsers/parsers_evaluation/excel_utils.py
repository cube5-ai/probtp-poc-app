"""
Excel helpers to write sheets and add charts for the evaluation report.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList


def write_excel_with_charts(
    global_metrics: dict[str, Any],
    pipeline_metrics: pd.DataFrame,
    detailed_results: pd.DataFrame,
    file_metrics: pd.DataFrame,
    pipeline_file_metrics: pd.DataFrame,
    output_file: Path,
):
    """Write Excel report with multiple sheets and charts (French labels)."""
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # 1. Synthèse globale
        global_df = pd.DataFrame([global_metrics])
        global_df.to_excel(writer, sheet_name="Synthèse globale", index=False)

        # 2. Comparaison des pipelines
        pipeline_metrics.to_excel(writer, sheet_name="Comparaison pipelines", index=False)

        # 3. Résultats détaillés
        detailed_results.to_excel(writer, sheet_name="Résultats détaillés", index=False)

        # 4. Répartition des verdicts
        verdict_pivot = detailed_results.pivot_table(
            index="pipeline",
            columns="verdict",
            aggfunc="size",
            fill_value=0,
        )
        verdict_pivot.to_excel(writer, sheet_name="Répartition des verdicts")

        # 5. Comparaison des fichiers
        if not file_metrics.empty:
            file_metrics.to_excel(writer, sheet_name="Comparaison fichiers", index=False)

        # 6. Pipeline x Fichier (taux de succès)
        if not pipeline_file_metrics.empty:
            pivot = pipeline_file_metrics.pivot(
                index="pipeline", columns="file_name", values="success_rate"
            ).reset_index()
            pivot.to_excel(writer, sheet_name="Pipeline x Fichier", index=False)

    # Add charts using openpyxl
    wb = load_workbook(output_file)

    add_global_pie_chart(wb, global_metrics)
    add_pipeline_bar_chart(wb, pipeline_metrics)

    if "Comparaison fichiers" in wb.sheetnames and not file_metrics.empty:
        add_file_bar_chart(wb, file_metrics)

    if "Pipeline x Fichier" in wb.sheetnames and not pipeline_file_metrics.empty:
        add_pipeline_file_grouped_bar_chart(wb)

    wb.save(output_file)


def add_global_pie_chart(wb, global_metrics: dict[str, Any]):
    """Add a pie chart showing overall verdict distribution (French labels)."""
    ws = wb["Synthèse globale"]

    chart_data = [
        ["Verdict", "Nombre"],
        ["Correct", global_metrics.get("correct_count", 0)],
        ["Partiel", global_metrics.get("partial_count", 0)],
        ["Hallucination", global_metrics.get("hallucination_count", 0)],
        ["Manqué", global_metrics.get("missed_count", 0)],
    ]

    start_row = ws.max_row + 3
    for i, row_data in enumerate(chart_data, start=start_row):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)

    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + 4)
    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + 4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Répartition globale des verdicts"

    ws.add_chart(pie, f"E{start_row}")


def add_pipeline_bar_chart(wb, pipeline_metrics: pd.DataFrame):
    """Add a bar chart comparing success rates across pipelines (French labels)."""
    ws = wb["Comparaison pipelines"]
    num_pipelines = len(pipeline_metrics)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Taux de succès par pipeline"
    chart.y_axis.title = "Taux de succès"
    chart.x_axis.title = "Pipeline"

    data = Reference(ws, min_col=ws.max_column, min_row=1, max_row=num_pipelines + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=num_pipelines + 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    ws.add_chart(chart, f"A{num_pipelines + 5}")


def add_file_bar_chart(wb, file_metrics: pd.DataFrame):
    """Add a bar chart comparing success rates across files (French labels)."""
    ws = wb["Comparaison fichiers"]
    num_files = len(file_metrics)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Taux de succès par fichier"
    chart.y_axis.title = "Taux de succès"
    chart.x_axis.title = "Fichier"

    data = Reference(ws, min_col=ws.max_column, min_row=1, max_row=num_files + 1)
    cats = Reference(ws, min_col=2, min_row=2, max_row=num_files + 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    ws.add_chart(chart, f"A{num_files + 5}")


def add_pipeline_file_grouped_bar_chart(wb):
    """Add a grouped bar chart showing success rates for each file per pipeline (French labels)."""
    ws = wb["Pipeline x Fichier"]
    num_rows = ws.max_row
    num_cols = ws.max_column

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.overlap = 0
    chart.title = "Taux de succès par pipeline et fichier"
    chart.y_axis.title = "Taux de succès"
    chart.x_axis.title = "Pipeline"

    data = Reference(ws, min_col=2, min_row=1, max_col=num_cols, max_row=num_rows)
    cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    ws.add_chart(chart, f"A{num_rows + 3}")


