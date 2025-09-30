"""
Generate comprehensive evaluation report with metrics, Excel dashboard, and AI analysis.
"""
import asyncio
import json
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from dotenv import load_dotenv
from google import genai
from google.genai.types import GenerateContentConfig, ThinkingConfig
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

load_dotenv()


# File name mappings from evaluate_pipelines.py
QA_FILE_FILE_NAME_MAP = {
    "#1": "File #1 - Laurent M - Conditions particulières - Socle AXA - SAE.md",
    "#2": "File #2 - Laurent M - tableau garantie fm 2025 word.md",
}


def load_evaluation_results(eval_dir: Path) -> list[dict[str, Any]]:
    """Load all evaluation results from JSON files in the eval directory."""
    all_results = []
    for json_file in eval_dir.glob("*.json"):
        with open(json_file, encoding="utf-8") as f:
            results = json.load(f)
            all_results.extend(results)
    return all_results


def compute_global_metrics(results: list[dict[str, Any]]) -> dict[str, Any]:
    """Compute overall performance metrics across all evaluations."""
    total = len(results)
    if total == 0:
        return {}

    verdict_counts = {
        "correct": sum(1 for r in results if r["verdict"] == "CORRECT"),
        "partial": sum(1 for r in results if r["verdict"] == "PARTIAL"),
        "hallucination": sum(1 for r in results if r["verdict"] == "HALLUCINATION"),
        "missed": sum(1 for r in results if r["verdict"] == "MISSED"),
    }

    avg_confidence = sum(r.get("confidence", 0.0) for r in results) / total

    return {
        "total_evaluations": total,
        "correct_count": verdict_counts["correct"],
        "partial_count": verdict_counts["partial"],
        "hallucination_count": verdict_counts["hallucination"],
        "missed_count": verdict_counts["missed"],
        "correct_rate": verdict_counts["correct"] / total,
        "partial_rate": verdict_counts["partial"] / total,
        "hallucination_rate": verdict_counts["hallucination"] / total,
        "missed_rate": verdict_counts["missed"] / total,
        "avg_confidence": avg_confidence,
        "success_rate": (verdict_counts["correct"] + verdict_counts["partial"]) / total,
    }


def compute_pipeline_metrics(results: list[dict[str, Any]]) -> pd.DataFrame:
    """Compute performance metrics grouped by pipeline."""
    df = pd.DataFrame(results)

    pipeline_stats = df.groupby("pipeline_name").agg({
        "verdict": [
            ("total", "count"),
            ("correct", lambda x: (x == "CORRECT").sum()),
            ("partial", lambda x: (x == "PARTIAL").sum()),
            ("hallucination", lambda x: (x == "HALLUCINATION").sum()),
            ("missed", lambda x: (x == "MISSED").sum()),
        ],
        "confidence": "mean"
    }).round(4)

    pipeline_stats.columns = ["_".join(col).strip("_") for col in pipeline_stats.columns]

    # Calculate rates
    pipeline_stats["correct_rate"] = (
        pipeline_stats["verdict_correct"] / pipeline_stats["verdict_total"]
    ).round(4)
    pipeline_stats["partial_rate"] = (
        pipeline_stats["verdict_partial"] / pipeline_stats["verdict_total"]
    ).round(4)
    pipeline_stats["hallucination_rate"] = (
        pipeline_stats["verdict_hallucination"] / pipeline_stats["verdict_total"]
    ).round(4)
    pipeline_stats["missed_rate"] = (
        pipeline_stats["verdict_missed"] / pipeline_stats["verdict_total"]
    ).round(4)
    pipeline_stats["success_rate"] = (
        (pipeline_stats["verdict_correct"] + pipeline_stats["verdict_partial"])
        / pipeline_stats["verdict_total"]
    ).round(4)

    return pipeline_stats.reset_index()


def build_question_to_file_map(eval_data_path: Path) -> dict[str, dict[str, str]]:
    """Build mapping from generated question text to file identifiers from eval_data.

    Returns mapping: question -> {"file_id": "#1", "file_name": "Pretty.md"}
    """
    if not eval_data_path.exists():
        return {}

    with open(eval_data_path, encoding="utf-8") as f:
        eval_data = json.load(f)

    mapping: dict[str, dict[str, str]] = {}
    for obj in eval_data:
        # Rebuild question exactly as used during evaluation
        question = "Quel est la prise en charge ou le remboursement pour"
        question += f" {obj['Catégorie 1']} > {obj['Catégorie 2']}"
        question += f" > {obj['Catégorie 3']}" if obj.get('Catégorie 3') else ""
        question += f" au niveau de garantie {obj['Niveau']} ?"

        file_id = obj.get("Fichier", "")
        file_name = QA_FILE_FILE_NAME_MAP.get(file_id, file_id)
        mapping[question] = {"file_id": file_id, "file_name": file_name}
    return mapping


def create_detailed_results_df(
    results: list[dict[str, Any]],
    question_to_file_map: dict[str, dict[str, str]] | None = None,
) -> pd.DataFrame:
    """Create a detailed DataFrame with all evaluation results, enriched with file info."""
    rows: list[dict[str, Any]] = []
    for r in results:
        question = r["question"]
        file_id = ""
        file_name = ""
        if question_to_file_map and question in question_to_file_map:
            file_id = question_to_file_map[question]["file_id"]
            file_name = question_to_file_map[question]["file_name"]

        rows.append({
            "pipeline": r["pipeline_name"],
            "file_id": file_id,
            "file_name": file_name,
            "question": question,
            "verdict": r["verdict"],
            "confidence": r.get("confidence", 0.0),
            "ground_truth": r["ground_truth"],
            "llm_answer": r["llm_answer"],
            "explanation": r["explanation"],
            "missing_info": "; ".join(r.get("missing_information", [])),
        })

    return pd.DataFrame(rows)


def compute_file_metrics(detailed_results: pd.DataFrame) -> pd.DataFrame:
    """Compute performance metrics grouped by file (across all pipelines)."""
    if detailed_results.empty:
        return pd.DataFrame()

    df = detailed_results.copy()
    group = df.groupby(["file_id", "file_name"])  # file_name may be empty if mapping not found
    stats = group.agg({
        "verdict": [
            ("total", "count"),
            ("correct", lambda x: (x == "CORRECT").sum()),
            ("partial", lambda x: (x == "PARTIAL").sum()),
            ("hallucination", lambda x: (x == "HALLUCINATION").sum()),
            ("missed", lambda x: (x == "MISSED").sum()),
        ],
        "confidence": "mean",
    }).round(4)

    stats.columns = ["_".join(col).strip("_") for col in stats.columns]
    stats["correct_rate"] = (stats["verdict_correct"] / stats["verdict_total"]).round(4)
    stats["partial_rate"] = (stats["verdict_partial"] / stats["verdict_total"]).round(4)
    stats["hallucination_rate"] = (stats["verdict_hallucination"] / stats["verdict_total"]).round(4)
    stats["missed_rate"] = (stats["verdict_missed"] / stats["verdict_total"]).round(4)
    stats["success_rate"] = (
        (stats["verdict_correct"] + stats["verdict_partial"]) / stats["verdict_total"]
    ).round(4)

    return stats.reset_index()


def compute_pipeline_file_metrics(detailed_results: pd.DataFrame) -> pd.DataFrame:
    """Compute performance metrics grouped by pipeline and file."""
    if detailed_results.empty:
        return pd.DataFrame()

    df = detailed_results.copy()
    group = df.groupby(["pipeline", "file_id", "file_name"])  # long format
    stats = group.agg({
        "verdict": [
            ("total", "count"),
            ("correct", lambda x: (x == "CORRECT").sum()),
            ("partial", lambda x: (x == "PARTIAL").sum()),
            ("hallucination", lambda x: (x == "HALLUCINATION").sum()),
            ("missed", lambda x: (x == "MISSED").sum()),
        ],
    })
    stats.columns = ["_".join(col).strip("_") for col in stats.columns]
    stats["success_rate"] = (
        (stats["verdict_correct"] + stats["verdict_partial"]) / stats["verdict_total"]
    ).round(4)
    return stats.reset_index()


def write_excel_with_charts(
    global_metrics: dict[str, Any],
    pipeline_metrics: pd.DataFrame,
    detailed_results: pd.DataFrame,
    file_metrics: pd.DataFrame,
    pipeline_file_metrics: pd.DataFrame,
    output_file: Path,
):
    """Write comprehensive Excel report with multiple sheets and charts."""
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Sheet 1: Global Summary
        global_df = pd.DataFrame([global_metrics])
        global_df.to_excel(writer, sheet_name="Global Summary", index=False)

        # Sheet 2: Pipeline Comparison
        pipeline_metrics.to_excel(writer, sheet_name="Pipeline Comparison", index=False)

        # Sheet 3: Detailed Results
        detailed_results.to_excel(writer, sheet_name="Detailed Results", index=False)

        # Sheet 4: Verdict Distribution by Pipeline
        verdict_pivot = detailed_results.pivot_table(
            index="pipeline",
            columns="verdict",
            aggfunc="size",
            fill_value=0
        )
        verdict_pivot.to_excel(writer, sheet_name="Verdict Distribution")

        # Sheet 5: File Comparison (aggregated across pipelines)
        if not file_metrics.empty:
            file_metrics.to_excel(writer, sheet_name="File Comparison", index=False)

        # Sheet 6: Pipeline x File (success rate pivot)
        if not pipeline_file_metrics.empty:
            pivot = pipeline_file_metrics.pivot(
                index="pipeline", columns="file_name", values="success_rate"
            ).reset_index()
            pivot.to_excel(writer, sheet_name="Pipeline x File", index=False)

    # Add charts using openpyxl
    wb = load_workbook(output_file)

    # Add pie chart to Global Summary
    add_global_pie_chart(wb, global_metrics)

    # Add bar chart to Pipeline Comparison
    add_pipeline_bar_chart(wb, pipeline_metrics)

    # Add bar chart to File Comparison
    if "File Comparison" in wb.sheetnames and not file_metrics.empty:
        add_file_bar_chart(wb, file_metrics)

    # Add grouped bar chart to Pipeline x File
    if "Pipeline x File" in wb.sheetnames and not pipeline_file_metrics.empty:
        add_pipeline_file_grouped_bar_chart(wb)

    wb.save(output_file)


def add_global_pie_chart(wb, global_metrics: dict[str, Any]):
    """Add a pie chart showing overall verdict distribution."""
    ws = wb["Global Summary"]

    # Create a small data table for the chart
    chart_data = [
        ["Verdict", "Count"],
        ["Correct", global_metrics["correct_count"]],
        ["Partial", global_metrics["partial_count"]],
        ["Hallucination", global_metrics["hallucination_count"]],
        ["Missed", global_metrics["missed_count"]],
    ]

    start_row = ws.max_row + 3
    for i, row_data in enumerate(chart_data, start=start_row):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)

    # Create pie chart
    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=start_row+1, max_row=start_row+4)
    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row+4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Overall Verdict Distribution"

    ws.add_chart(pie, f"E{start_row}")


def add_pipeline_bar_chart(wb, pipeline_metrics: pd.DataFrame):
    """Add a bar chart comparing success rates across pipelines."""
    ws = wb["Pipeline Comparison"]

    # Find the success_rate column
    num_pipelines = len(pipeline_metrics)

    # Create bar chart for success rates
    chart = BarChart()
    chart.type = "col"
    chart.title = "Success Rate by Pipeline"
    chart.y_axis.title = "Success Rate"
    chart.x_axis.title = "Pipeline"

    # Assuming pipeline_name is in column 1 and success_rate is in the last column
    data = Reference(ws, min_col=ws.max_column, min_row=1, max_row=num_pipelines+1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=num_pipelines+1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    ws.add_chart(chart, f"A{num_pipelines + 5}")


def add_file_bar_chart(wb, file_metrics: pd.DataFrame):
    """Add a bar chart comparing success rates across files."""
    ws = wb["File Comparison"]

    # Determine rows: header + rows
    num_files = len(file_metrics)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Success Rate by File"
    chart.y_axis.title = "Success Rate"
    chart.x_axis.title = "File"

    # Assume columns: file_id, file_name, ..., success_rate is last
    data = Reference(ws, min_col=ws.max_column, min_row=1, max_row=num_files + 1)
    cats = Reference(ws, min_col=2, min_row=2, max_row=num_files + 1)  # file_name

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    ws.add_chart(chart, f"A{num_files + 5}")


def add_pipeline_file_grouped_bar_chart(wb):
    """Add a grouped bar chart showing success rates for each file per pipeline."""
    ws = wb["Pipeline x File"]

    # Determine size
    num_rows = ws.max_row
    num_cols = ws.max_column

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.overlap = 0
    chart.title = "Success Rate by Pipeline and File"
    chart.y_axis.title = "Success Rate"
    chart.x_axis.title = "Pipeline"

    # First column is pipeline (categories), data series are columns 2..num_cols
    data = Reference(ws, min_col=2, min_row=1, max_col=num_cols, max_row=num_rows)
    cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    ws.add_chart(chart, f"A{num_rows + 3}")


async def generate_ai_report(
    global_metrics: dict[str, Any],
    pipeline_metrics: pd.DataFrame,
    detailed_results: pd.DataFrame,
    file_metrics: pd.DataFrame,
    pipeline_file_metrics: pd.DataFrame,
    pipeline_summaries_md: str = "",
) -> str:
    """Generate comprehensive analysis report using Gemini with extended thinking (French)."""

    prompt = f"""
Tu es un expert data scientist analysant la performance de différents pipelines d'extraction de documents.

## Méthodologie d'évaluation (description factuelle)

Décris succinctement les étapes suivies sans les juger ni les critiquer. Nous avons évalué plusieurs pipelines en:
1. Parsant des documents d'assurance (PDF) avec différentes méthodes
2. Posant des questions ciblées sur la prise en charge et le remboursement
3. Comparant les réponses du LLM (basées sur les documents parsés) au jeu de vérité terrain
4. Attribuant un verdict selon 4 catégories:
   - CORRECT: Concordance parfaite avec la vérité terrain
   - PARTIAL: Bonne direction mais informations clés manquantes
   - HALLUCINATION: Informations incorrectes ou non étayées
   - MISSED: Indique "non disponible" alors que l'information existe

   Note: les pipelines incluent des approches « parsing + QA » et des approches VQA (question-réponse visuelle) qui interrogent directement les PDF sans parsing préalable.

## Pipelines évalués (synthèse)

{pipeline_summaries_md if pipeline_summaries_md else "(Synthèse non disponible)"}

## Indicateurs globaux

{json.dumps(global_metrics, indent=2, ensure_ascii=False)}

## Comparaison des pipelines

{pipeline_metrics.to_string()}

## Performance par fichier

{file_metrics.to_string()}

## Pipeline x Fichier (taux de succès)

{pipeline_file_metrics.pivot(index='pipeline', columns='file_name', values='success_rate').to_string() if not pipeline_file_metrics.empty else "Aucune donnée pipeline x fichier"}

## Exemples de résultats détaillés (10 premiers)

{detailed_results.head(10).to_string()}

## Ta mission

Rédige un rapport d'analyse structuré en français, incluant:

1. Résumé exécutif: constats clés et recommandations (2-3 paragraphes)

2. Description de la méthodologie: reformulation neutre et factuelle (sans évaluation)

3. **Description détaillée des pipelines évalués:**
   Pour chaque pipeline listé dans "Pipelines évalués (synthèse)" ci-dessus, crée une sous-section avec:
   - Nom du pipeline (titre de niveau 3)
   - Approche technique: parsing utilisé, modèles/APIs, étapes clés
   - Points forts théoriques de l'approche
   - Limitations potentielles
   Utilise les informations fournies dans la section "Pipelines évalués (synthèse)" et synthétise-les de manière claire et structurée.

4. Analyse de performance:
   - Comparaison objective des pipelines
   - Motifs récurrents de réussite/échec
   - Corrélation entre score de confiance et verdict

5. Analyse des causes:
   - Hypothèses sur les écarts de performance entre pipelines (en lien avec leurs approches techniques décrites)
   - Caractéristiques documentaires influentes (structure, tableaux, formatage)

6. Recommandations:
   - Pipeline(s) à privilégier et justification
   - Idées d'approches hybrides
   - Pistes d'amélioration de l'évaluation

7. Analyse par type de verdict:
   - PARTIAL: informations manquantes typiques
   - MISSED: angles morts systématiques éventuels
   - HALLUCINATION: sources fréquentes d'erreurs

8. Aperçus par fichier:
   - Variabilité entre fichiers et explications plausibles

Utilise un style clair, structuré en markdown, avec des listes lorsque pertinent.
"""

    client = genai.Client(
        vertexai=True,
        project="probtp-poc-prod",
        location="global",
    )

    try:
        response = await client.aio.models.generate_content(
            model="gemini-2.5-flash-preview-09-2025",
            contents=prompt,
            config=GenerateContentConfig(
                thinking_config=ThinkingConfig(
                    thinking_budget=8192  # Extended reasoning budget
                )
            ),
        )

        report = getattr(response, "text", "").strip()
        return report

    finally:
        if hasattr(client, "close"):
            if asyncio.iscoroutinefunction(client.close):
                await client.close()
            else:
                client.close()


async def generate_full_report():
    """Main function to generate complete evaluation report."""
    print("Chargement des résultats d'évaluation...")
    eval_dir = Path("output/evals")
    results = load_evaluation_results(eval_dir)
    print(f"{len(results)} résultats chargés")

    print("\nCalcul des métriques...")
    # Build mapping question -> file
    q_to_file = build_question_to_file_map(Path("eval_data/data_0.json"))

    global_metrics = compute_global_metrics(results)
    pipeline_metrics = compute_pipeline_metrics(results)
    detailed_results = create_detailed_results_df(results, q_to_file)
    file_metrics = compute_file_metrics(detailed_results)
    pipeline_file_metrics = compute_pipeline_file_metrics(detailed_results)

    print("\nMétriques globales:")
    print(json.dumps(global_metrics, indent=2))

    print("\nMétriques par pipeline:")
    print(pipeline_metrics.to_string())

    if not file_metrics.empty:
        print("\nMétriques par fichier:")
        print(file_metrics.to_string())

    if not pipeline_file_metrics.empty:
        print("\nPipeline x Fichier (taux de succès):")
        print(pipeline_file_metrics.pivot(index='pipeline', columns='file_name', values='success_rate').to_string())

    # Generate Excel report
    output_dir = Path("reports")
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = output_dir / f"evaluation_report_{timestamp}.xlsx"

    print(f"\nGénération du rapport Excel: {excel_file}")
    write_excel_with_charts(
        global_metrics,
        pipeline_metrics,
        detailed_results,
        file_metrics,
        pipeline_file_metrics,
        excel_file,
    )
    print(f"✓ Rapport Excel enregistré: {excel_file}")

    # Generate AI analysis report (with pipeline summaries)
    print("\nGénération du rapport d'analyse IA (cela peut prendre une minute)...")
    # Build pipeline summaries from scripts in parsing_pipelines/
    try:
        from .ai_report_utils import summarize_pipeline_files_fr
    except Exception:
        # Fallback to absolute import if package context differs
        from ai_report_utils import summarize_pipeline_files_fr

    pipelines_dir = Path("parsing_pipelines")
    vqa_dir = Path("vqa_pipelines")
    pipeline_summaries = await summarize_pipeline_files_fr(pipelines_dir)
    try:
        from .ai_report_utils import summarize_vqa_pipelines_fr
    except Exception:
        from ai_report_utils import summarize_vqa_pipelines_fr
    vqa_summaries = await summarize_vqa_pipelines_fr(vqa_dir)

    sections = []
    if pipeline_summaries:
        sections.append("\n\n".join([f"### {name}\n\n{summary}" for name, summary in pipeline_summaries.items()]))
    if vqa_summaries:
        sections.append("\n\n".join([f"### {name} (VQA)\n\n{summary}" for name, summary in vqa_summaries.items()]))
    pipeline_summaries_md = "\n\n".join(sections)

    ai_report = await generate_ai_report(
        global_metrics,
        pipeline_metrics,
        detailed_results,
        file_metrics,
        pipeline_file_metrics,
        pipeline_summaries_md,
    )

    report_file = output_dir / f"ai_analysis_report_{timestamp}.md"
    with open(report_file, "w", encoding="utf-8") as f:
        f.write("# Rapport d'évaluation des pipelines d'extraction\n\n")
        f.write(f"**Généré :** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write("---\n\n")
        f.write(ai_report)

    print(f"✓ Rapport IA enregistré: {report_file}")

    return excel_file, report_file


if __name__ == "__main__":
    excel_path, report_path = asyncio.run(generate_full_report())
    print(f"\n{'='*60}")
    print("Génération du rapport terminée !")
    print(f"Tableau de bord Excel: {excel_path}")
    print(f"Analyse IA: {report_path}")
    print(f"{'='*60}")
